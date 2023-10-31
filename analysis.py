import copy
from typing import Dict, Optional, Union

import pandas as pd
from openpyxl.cell import Cell
from openpyxl.utils import column_index_from_string
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.worksheet import Worksheet

import analyzers
from table_stats import TableStats


def strip_prefix(string: str, prefix: str) -> str:
    """Strip prefix from string if it exists"""
    if string.startswith(prefix):
        return string[len(prefix) :]
    return string


# ---------------------------------------------
#   Helpers for openpyxl
# ---------------------------------------------


def letter_to_index(c: Union[str, int]) -> int:
    """Convert a column index or column name (eg. J, AA) to a column index."""
    if isinstance(c, str):
        return column_index_from_string(c)
    return c


def copy_columns(ws: Worksheet, column_range: CellRange, dest_column: Union[str, int]) -> None:
    """Copy and paste columns with `column_range` to the columns starting fromm `dest_column`."""
    dest_column = letter_to_index(dest_column)

    # copy cell data and formatting
    for column in ws.iter_cols(column_range.min_col, column_range.max_col):
        for cell in column:
            new_cell = ws.cell(
                row=cell.row, column=dest_column + (cell.column - column_range.min_col), value=cell.value
            )
            if cell.has_style:
                new_cell.font = copy.copy(cell.font)
                new_cell.border = copy.copy(cell.border)
                new_cell.fill = copy.copy(cell.fill)
                new_cell.number_format = copy.copy(cell.number_format)
                new_cell.protection = copy.copy(cell.protection)
                new_cell.alignment = copy.copy(cell.alignment)

    # copy merged cells
    for mcr in set(ws.merged_cells.ranges):
        if not (column_range.min_col <= mcr.min_col <= column_range.max_col):
            continue
        cr = CellRange(mcr.coord)
        cr.shift(col_shift=dest_column - column_range.min_col)
        ws.merge_cells(cr.coord)


def write_table(df: pd.DataFrame, cell: Cell, index: bool = False, header: bool = False) -> None:
    """Write Pandas dataframe to spreadsheet, starting from cell and going down and right"""
    for dr, row in enumerate(dataframe_to_rows(df, index=index, header=header)):
        for dc, value in enumerate(row):
            cell.offset(dr, dc).value = value


def write_row(row: list, cell: Cell) -> None:
    """Write list of values to a spreadsheet, starting from a cell and moving right."""
    for dc, value in enumerate(row):
        cell.offset(0, dc).value = value


# ---------------------------------------------
#   Table generation code
# ---------------------------------------------


def create_general_sheet(ws: Worksheet, stats: TableStats, mode_labels: Optional[Dict[str, str]] = None) -> None:
    """Create General sheet"""
    if mode_labels is None:
        mode_labels = {"dance-single": "Singles", "dance-double": "Doubles"}

    a = analyzers.chart_counts_for_each_pack(stats, mode_labels)
    write_table(a.reset_index(), ws["A4"])

    ws["A3"].value = len(a)  # set pack count
    write_row(list(a.sum()), ws["B3"])  # set song and chart totals for each mode

    # render difficulty histogram
    # todo: make sure both tables display packs in the right order
    a = analyzers.pack_difficulty_histogram(stats)
    write_table(a.reset_index().drop("pack", axis="columns"), ws["H3"], header=True)


def create_most_played_charts_sheet(ws: Worksheet, stats: TableStats, limit: int = 50) -> None:
    """Create Most Played Charts sheet"""
    all_songs = analyzers.most_played_charts(stats, limit, modes=["dance-single", "dance-double"])
    doubles_only = analyzers.most_played_charts(stats, limit, modes=["dance-double"])

    write_table(all_songs, ws["A3"])
    write_table(doubles_only, ws["I3"])

    # todo: set background colour for extra song entries?


def create_most_played_songs_sheet(ws: Worksheet, stats: TableStats, limit: int = 50) -> None:
    """Create Most Played Songs sheet"""
    all_songs = analyzers.most_played_songs(stats, limit, modes=["dance-single", "dance-double"])
    doubles_only = analyzers.most_played_songs(stats, limit, modes=["dance-double"])

    write_table(all_songs, ws["A3"])
    write_table(doubles_only, ws["A57"])

    # todo: set background colour for extra song entries?
    # todo: doesn't work for limits > 50, decide what to do


def create_most_played_packs_sheet(ws: Worksheet, stats: TableStats) -> None:
    """Create Most Played Packs sheet"""
    packs_by_playcount = analyzers.most_played_packs(stats)
    song_breakdown = analyzers.most_played_charts_per_pack(stats)

    final_table = packs_by_playcount.join(song_breakdown)
    write_table(final_table.reset_index(), ws["A2"])


def create_recently_played_packs_sheet(ws: Worksheet, stats: TableStats) -> None:
    """Create Recently Played Packs sheet"""
    packs = analyzers.recently_played_packs(stats)
    write_table(packs.reset_index(), ws["A2"])


def create_pack_completion_sheet(ws: Worksheet, stats: TableStats) -> None:
    """Create Pack Completion sheet"""
    completion = analyzers.pack_completion(stats)

    # note: column labels for grade boundaries aren't written out in code,
    # they're hardcoded in the Excel sheet so user has maximal control over formatting.
    # if grade boundaries are changed the excel sheet has to be modified as well
    grade_breakdown = analyzers.pack_score_breakdown(stats, analyzers.GRADE_BY_10)

    table = completion.join(grade_breakdown)
    write_table(table.reset_index(), ws["A3"])


def create_highest_scores_sheet(ws: Worksheet, stats: TableStats) -> None:
    """Create Highest Scores + Passes sheet"""
    # ideas for other ways to split it
    #   - top 5 for each block difficulty
    #   - highest scores (DDR only)
    # there are too many ways to slice the highscore data --
    # I think the only way to make it useful is to make it interactive
    a = analyzers.song_grades_by_meter(stats, analyzers.GRADE_BY_10)
    highest_scores = analyzers.highest_scores(stats, with_ddr=False)
    highest_passes = analyzers.highest_passes(stats, with_ddr=False)

    write_table(a.reset_index(), ws["S2"])
    write_table(highest_scores, ws["A29"])
    write_table(highest_passes, ws["J29"])

    # todo: doubles sheet
